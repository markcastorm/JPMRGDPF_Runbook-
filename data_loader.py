"""
JPMRGDPF Data Loader
====================
Loads input Excel files and locates the target worksheet.

This module handles:
- Scanning the input directory for Excel/CSV files
- Opening workbooks and finding the "Global Outlook Summary" tab (flexible matching)
- Returning raw worksheet data for the parser to analyze dynamically
"""

import os
import logging
from glob import glob

import openpyxl

import config

logger = logging.getLogger(__name__)


class JPMRGDPFDataLoader:
    """Loads JPM Global Outlook Excel files."""

    def __init__(self):
        self.input_dir = config.INPUT_DIR

    def find_input_files(self):
        """
        Find all input files in the input directory.

        Returns:
            list: Sorted list of file paths (newest first by modification time).
        """
        patterns = ['*.xlsx', '*.xlsm', '*.xls']
        files = []

        for pattern in patterns:
            search_path = os.path.join(self.input_dir, pattern)
            found = glob(search_path)
            # Skip temp files (start with ~$)
            found = [f for f in found if not os.path.basename(f).startswith('~$')]
            files.extend(found)

        files.sort(key=os.path.getmtime, reverse=True)

        logger.info(f'Found {len(files)} input file(s) in {self.input_dir}')
        for f in files:
            logger.debug(f'  - {os.path.basename(f)}')

        return files

    def _find_target_sheet(self, wb):
        """
        Find the target worksheet by flexible name matching.

        Searches for sheet names that contain any of the TARGET_TAB_NAMES
        (case-insensitive, partial match).

        Args:
            wb: openpyxl Workbook object.

        Returns:
            Worksheet object or None.
        """
        available_sheets = wb.sheetnames
        logger.info(f'Available sheets: {available_sheets}')

        for target_name in config.TARGET_TAB_NAMES:
            # Exact match first
            if target_name in available_sheets:
                logger.info(f'Found exact match: "{target_name}"')
                return wb[target_name]

            # Case-insensitive match
            for sheet_name in available_sheets:
                if sheet_name.lower() == target_name.lower():
                    logger.info(f'Found case-insensitive match: "{sheet_name}"')
                    return wb[sheet_name]

            # Partial match (sheet name contains target)
            for sheet_name in available_sheets:
                if target_name.lower() in sheet_name.lower():
                    logger.info(f'Found partial match: "{sheet_name}" (contains "{target_name}")')
                    return wb[sheet_name]

        logger.warning(f'No matching sheet found. Available: {available_sheets}')
        return None

    def _read_sheet_data(self, ws):
        """
        Read all cell values from a worksheet into a 2D list.

        Returns:
            list[list]: Grid of cell values indexed as data[row][col] (0-based).
        """
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column,
                                values_only=True):
            data.append(list(row))

        logger.info(f'Read {len(data)} rows x {ws.max_column} cols from sheet')
        return data

    def load_file(self, file_path):
        """
        Load a single input file and return raw sheet data.

        Args:
            file_path: Path to the input Excel file.

        Returns:
            dict with:
                - 'file_name': basename of the file
                - 'file_path': full path
                - 'sheet_name': name of the matched sheet
                - 'data': 2D list of all cell values (0-based row/col)
                - 'max_rows': number of rows
                - 'max_cols': number of columns
            Returns None if no matching sheet found.
        """
        logger.info(f'Loading file: {os.path.basename(file_path)}')

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            ws = self._find_target_sheet(wb)
            if ws is None:
                logger.error(f'No target sheet found in {file_path}')
                wb.close()
                return None

            sheet_name = ws.title
            data = self._read_sheet_data(ws)
            max_rows = len(data)
            max_cols = len(data[0]) if data else 0

            wb.close()

            logger.info(f'Loaded "{sheet_name}" from {os.path.basename(file_path)} '
                        f'({max_rows} rows x {max_cols} cols)')

            return {
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'sheet_name': sheet_name,
                'data': data,
                'max_rows': max_rows,
                'max_cols': max_cols,
            }

        except Exception as e:
            logger.error(f'Error loading file {file_path}: {e}')
            raise

    def load_all_files(self):
        """
        Load all input files.

        Returns:
            list: List of loaded data dictionaries (newest first).
        """
        files = self.find_input_files()
        results = []

        for file_path in files:
            try:
                data = self.load_file(file_path)
                if data:
                    results.append(data)
            except Exception as e:
                logger.error(f'Failed to load {file_path}: {e}')
                if not config.CONTINUE_ON_ERROR:
                    raise

        return results


if __name__ == '__main__':
    from logger_setup import setup_logging
    setup_logging()

    loader = JPMRGDPFDataLoader()
    files = loader.find_input_files()
    print(f'Found {len(files)} input files')

    if files:
        result = loader.load_file(files[0])
        if result:
            print(f"Loaded: {result['file_name']}")
            print(f"Sheet: {result['sheet_name']}")
            print(f"Size: {result['max_rows']} rows x {result['max_cols']} cols")
            print(f"\nFirst 5 rows (first 6 cols):")
            for i, row in enumerate(result['data'][:5]):
                print(f"  Row {i}: {[str(v)[:25] for v in row[:6]]}")
