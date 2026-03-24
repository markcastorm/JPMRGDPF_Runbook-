"""
JPMRGDPF Data Parser
====================
The brain of the pipeline. Dynamically scans the input sheet to locate
the "Real GDP % over year ago" section, identify year columns, match
target countries, and extract GDP forecast values.

Design philosophy: NEVER hardcode row/column positions. Always scan and
detect structure dynamically so the parser works even when the layout shifts.
"""

import os
import re
import logging

import pandas as pd
import openpyxl

import config

logger = logging.getLogger(__name__)


class JPMRGDPFParser:
    """Dynamically parses JPM Global Outlook data and merges with master."""

    def __init__(self):
        self.column_order = config.COLUMN_ORDER

    # =========================================================================
    # DYNAMIC SECTION DETECTION
    # =========================================================================

    def _find_section_header(self, data):
        """
        Scan the entire sheet to find the "Real GDP % over year ago" header.

        Searches every cell for text matching the target section header.
        Uses fuzzy matching (contains, case-insensitive) to handle variations.

        Args:
            data: 2D list of cell values (0-based).

        Returns:
            (row_index, col_index) of the header cell, or None if not found.
        """
        target = config.TARGET_SECTION_HEADER.lower()

        for row_idx, row in enumerate(data):
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip().lower()
                # Check for exact or partial match
                if target in cell_str:
                    logger.info(f'Found section header "{cell}" at row {row_idx}, col {col_idx}')
                    return (row_idx, col_idx)

        logger.error(f'Section header "{config.TARGET_SECTION_HEADER}" not found in sheet')
        return None

    def _find_year_columns(self, data, header_row, header_col):
        """
        Dynamically identify year columns below or adjacent to the section header.

        Scans rows near the header to find cells that look like years (4-digit numbers
        in a reasonable range like 2020-2040). Returns the year values and their
        column indices.

        Args:
            data: 2D list of cell values.
            header_row: Row index of the section header.
            header_col: Column index of the section header.

        Returns:
            list of (year, col_index) tuples sorted by column index.
        """
        years_found = []

        # Search in the header row itself and the next few rows
        search_rows = range(header_row, min(header_row + 3, len(data)))

        for row_idx in search_rows:
            row = data[row_idx]
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                # Try to detect year values
                year = self._extract_year(cell)
                if year and year not in [y for y, _ in years_found]:
                    years_found.append((year, col_idx))
                    logger.debug(f'Found year {year} at row {row_idx}, col {col_idx}')

        # Sort by column index (left to right)
        years_found.sort(key=lambda x: x[1])

        if years_found:
            year_labels = [str(y) for y, _ in years_found]
            logger.info(f'Detected {len(years_found)} year columns: {", ".join(year_labels)}')
        else:
            logger.error('No year columns detected near the section header')

        return years_found

    def _extract_year(self, value):
        """
        Try to extract a year from a cell value.

        Handles: int (2025), float (2025.0), string ("2025").

        Returns:
            int year or None.
        """
        if value is None:
            return None

        # Numeric
        if isinstance(value, (int, float)):
            val = int(value)
            if 2020 <= val <= 2040:
                return val
            return None

        # String
        text = str(value).strip()
        match = re.match(r'^(\d{4})$', text)
        if match:
            year = int(match.group(1))
            if 2020 <= year <= 2040:
                return year

        return None

    def _find_area_column(self, data, header_row):
        """
        Find the column that contains country/area names.

        Looks for the "Area" label in the header row or nearby rows.
        Falls back to checking which column has the most country name matches.

        Args:
            data: 2D list of cell values.
            header_row: Row index of the section header.

        Returns:
            int: Column index of the area column.
        """
        # Search near the header for "Area" label
        search_rows = range(max(0, header_row - 2), min(header_row + 3, len(data)))

        for row_idx in search_rows:
            row = data[row_idx]
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                if str(cell).strip().lower() == config.AREA_COLUMN_LABEL.lower():
                    logger.info(f'Found "{config.AREA_COLUMN_LABEL}" label at row {row_idx}, col {col_idx}')
                    return col_idx

        # Fallback: find column with most country name matches
        logger.warning('"Area" label not found, using heuristic detection')
        return self._detect_area_column_heuristic(data, header_row)

    def _detect_area_column_heuristic(self, data, header_row):
        """
        Heuristic: find which column contains the most recognized country names.

        Scans rows below the header and counts country matches per column.
        """
        # Only check first ~10 columns
        max_check_cols = min(10, len(data[0]) if data else 0)
        data_rows = data[header_row + 1: header_row + 50]

        best_col = 0
        best_count = 0

        for col_idx in range(max_check_cols):
            count = 0
            for row in data_rows:
                if col_idx < len(row) and row[col_idx]:
                    name = str(row[col_idx]).strip()
                    if config.is_target_country(name) or config.is_skip_area(name):
                        count += 1
            if count > best_count:
                best_count = count
                best_col = col_idx

        logger.info(f'Heuristic: area column = {best_col} ({best_count} matches)')
        return best_col

    def _find_data_start_row(self, data, header_row, area_col):
        """
        Find where actual data rows begin (after header/spacer rows).

        Scans down from the header looking for the first row that has a
        recognized country name or region in the area column.

        Args:
            data: 2D list of cell values.
            header_row: Row index of the section header.
            area_col: Column index of the area column.

        Returns:
            int: Row index where data starts.
        """
        for row_idx in range(header_row + 1, min(header_row + 10, len(data))):
            row = data[row_idx]
            if area_col < len(row) and row[area_col]:
                name = str(row[area_col]).strip()
                if name and len(name) > 1 and name.lower() != 'area':
                    logger.info(f'Data starts at row {row_idx} (first entry: "{name}")')
                    return row_idx

        # Default: header + 2 (skip one spacer row)
        logger.warning(f'Could not detect data start, defaulting to row {header_row + 2}')
        return header_row + 2

    # =========================================================================
    # DATA EXTRACTION
    # =========================================================================

    def extract_data(self, loaded_data):
        """
        Dynamically extract GDP forecast data from the loaded sheet.

        This is the main intelligence method. It:
        1. Finds the section header
        2. Identifies year columns
        3. Locates the area column
        4. Scans for target countries
        5. Extracts values for each country-year pair

        Args:
            loaded_data: Dictionary from data_loader with 'data' key.

        Returns:
            dict with:
                - 'years': list of year integers
                - 'country_data': dict mapping year -> {iso3_code: value}
                - 'file_name': input file name
                - 'countries_found': list of matched country names
                - 'countries_missing': list of expected but not found countries
        """
        data = loaded_data['data']
        file_name = loaded_data['file_name']

        logger.info(f'Parsing data from {file_name}')

        # Step 1: Find section header
        header_pos = self._find_section_header(data)
        if header_pos is None:
            raise ValueError(f'Cannot find "{config.TARGET_SECTION_HEADER}" in {file_name}')

        header_row, header_col = header_pos

        # Step 2: Find year columns
        year_columns = self._find_year_columns(data, header_row, header_col)
        if not year_columns:
            raise ValueError(f'No year columns found in {file_name}')

        years = [y for y, _ in year_columns]

        # Step 3: Find area column
        area_col = self._find_area_column(data, header_row)

        # Step 4: Find data start row
        data_start = self._find_data_start_row(data, header_row, area_col)

        # Step 5: Extract country data
        # Structure: {year: {iso3_code: value}}
        country_data = {year: {} for year in years}
        countries_found = []
        all_expected = {iso3 for iso3, _ in config.COUNTRY_ORDER}

        for row_idx in range(data_start, len(data)):
            row = data[row_idx]
            if area_col >= len(row):
                continue

            area_name = row[area_col]
            if area_name is None:
                continue
            area_name = str(area_name).strip()

            if not area_name:
                continue

            # Skip region aggregates
            if config.is_skip_area(area_name):
                logger.debug(f'Skipping region: {area_name}')
                continue

            # Check if this is a target country
            country_info = config.get_country_info(area_name)
            if country_info is None:
                logger.debug(f'Skipping non-target area: {area_name}')
                continue

            iso3_code, display_name = country_info

            # Extract values for each year
            for year, year_col in year_columns:
                if year_col < len(row):
                    raw_value = row[year_col]
                    cleaned = self._clean_value(raw_value)
                    country_data[year][iso3_code] = cleaned
                    logger.debug(f'  {display_name} [{iso3_code}] {year}: {cleaned}')

            countries_found.append(display_name)

        # Check for missing countries
        found_iso3 = set()
        for year_data in country_data.values():
            found_iso3.update(year_data.keys())

        missing = all_expected - found_iso3
        countries_missing = [name for iso3, name in config.COUNTRY_ORDER if iso3 in missing]

        logger.info(f'Extracted data for {len(countries_found)} countries across {len(years)} years')
        if countries_missing:
            logger.warning(f'Missing countries: {", ".join(countries_missing)}')

        return {
            'years': years,
            'country_data': country_data,
            'file_name': file_name,
            'countries_found': countries_found,
            'countries_missing': countries_missing,
        }

    def _clean_value(self, value):
        """
        Clean a cell value.

        Returns float, or NA_VALUE for missing data.
        """
        if value in config.NA_INPUT_VALUES:
            return config.NA_VALUE

        if isinstance(value, str):
            stripped = value.strip()
            if stripped in ['-', '--', '']:
                return config.NA_VALUE
            try:
                return float(stripped)
            except ValueError:
                return config.NA_VALUE

        if isinstance(value, (int, float)):
            return float(value)

        return config.NA_VALUE

    # =========================================================================
    # MASTER DATA MANAGEMENT
    # =========================================================================

    def load_master_data(self):
        """
        Load existing master data file.

        Returns:
            pd.DataFrame with 'date' column (years) and country code columns,
            or empty DataFrame if master doesn't exist.
        """
        if not os.path.exists(config.MASTER_DATA_FILE):
            logger.info('Master data file not found, will create new')
            return pd.DataFrame()

        try:
            logger.info(f'Loading master data from {config.MASTER_DATA_FILE}')

            wb = openpyxl.load_workbook(config.MASTER_DATA_FILE)
            ws = wb.active

            # Read column codes from row 1
            columns = ['date']
            for col in range(2, ws.max_column + 1):
                code = ws.cell(row=1, column=col).value
                if code:
                    columns.append(str(code))

            # Read data from row 3+
            data = []
            for row in range(3, ws.max_row + 1):
                row_data = []
                for col in range(1, len(columns) + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                if row_data[0] is not None:
                    data.append(row_data)

            wb.close()

            if not data:
                logger.info('Master file exists but has no data rows')
                return pd.DataFrame(columns=columns)

            df = pd.DataFrame(data, columns=columns)
            logger.info(f'Loaded {len(df)} rows from master data')
            return df

        except Exception as e:
            logger.error(f'Error loading master data: {e}')
            raise

    def build_dataframe(self, extracted):
        """
        Build a DataFrame from extracted data.

        Rows are years, columns are country codes (matching COLUMN_ORDER).

        Args:
            extracted: Dictionary from extract_data().

        Returns:
            pd.DataFrame
        """
        years = extracted['years']
        country_data = extracted['country_data']
        codes = [code for code, _ in self.column_order]

        rows = []
        for year in sorted(years):
            row = {'date': int(year)}
            year_data = country_data.get(year, {})
            for code, _ in self.column_order:
                # Extract iso3 from code: JPMRGDPF.REALGDP.ANNUAL.<ISO3>.A
                parts = code.split('.')
                iso3 = parts[3] if len(parts) >= 4 else None
                row[code] = year_data.get(iso3, config.NA_VALUE)
            rows.append(row)

        df = pd.DataFrame(rows, columns=['date'] + codes)
        logger.info(f'Built DataFrame: {len(df)} rows x {len(df.columns)} cols')
        return df

    def check_for_changes(self, master_df, new_df):
        """
        Compare new data against master to detect changes.

        Returns:
            dict with:
                - 'has_changes': bool
                - 'new_years': list of years not in master
                - 'updated_years': list of years with changed values
                - 'unchanged_years': list of years with no changes
        """
        if master_df.empty:
            return {
                'has_changes': True,
                'new_years': sorted(new_df['date'].tolist()),
                'updated_years': [],
                'unchanged_years': [],
            }

        master_years = set(master_df['date'].apply(lambda x: int(float(x)) if x else None).dropna())
        new_years_set = set(new_df['date'].apply(lambda x: int(float(x))))

        brand_new = sorted(new_years_set - master_years)
        common = sorted(new_years_set & master_years)

        updated = []
        unchanged = []

        data_cols = [c for c in new_df.columns if c != 'date']

        for year in common:
            master_row = master_df[master_df['date'].apply(lambda x: int(float(x))) == year]
            new_row = new_df[new_df['date'].apply(lambda x: int(float(x))) == year]

            if master_row.empty:
                updated.append(year)
                continue

            changed = False
            for col in data_cols:
                old_val = master_row.iloc[0].get(col)
                new_val = new_row.iloc[0].get(col)

                # Normalize for comparison
                old_str = str(old_val).strip() if old_val is not None else config.NA_VALUE
                new_str = str(new_val).strip() if new_val is not None else config.NA_VALUE

                if old_str != new_str:
                    changed = True
                    break

            if changed:
                updated.append(year)
            else:
                unchanged.append(year)

        has_changes = len(brand_new) > 0 or len(updated) > 0

        return {
            'has_changes': has_changes,
            'new_years': brand_new,
            'updated_years': updated,
            'unchanged_years': unchanged,
        }

    def merge_data(self, master_df, new_df):
        """
        Replace master with the latest rolling window of years.

        Per runbook: keep only the years present in the new input
        (last year, current year, future year). Old years that fall
        off the input are discarded from the master.

        Returns:
            pd.DataFrame: New data sorted by year.
        """
        new_years = sorted(new_df['date'].apply(lambda x: int(float(x))).tolist())

        if not master_df.empty:
            old_years = sorted(master_df['date'].apply(lambda x: int(float(x)) if x else 0).tolist())
            dropped = [y for y in old_years if y not in new_years]
            if dropped:
                logger.info(f'Dropping old years no longer in input: {dropped}')

        logger.info(f'Master rolling window: {new_years}')

        combined = new_df.copy()
        combined = combined.sort_values('date').reset_index(drop=True)

        logger.info(f'Master will have {len(combined)} rows')
        return combined

    def get_column_order(self):
        """Return the column order list of (code, description) tuples."""
        return self.column_order


if __name__ == '__main__':
    from logger_setup import setup_logging
    from data_loader import JPMRGDPFDataLoader

    setup_logging()

    loader = JPMRGDPFDataLoader()
    files = loader.find_input_files()

    if files:
        loaded = loader.load_file(files[0])
        if loaded:
            parser = JPMRGDPFParser()
            extracted = parser.extract_data(loaded)

            print(f"\nYears: {extracted['years']}")
            print(f"Countries found: {len(extracted['countries_found'])}")
            print(f"Countries missing: {extracted['countries_missing']}")

            df = parser.build_dataframe(extracted)
            print(f"\nDataFrame shape: {df.shape}")
            print(df.to_string(index=False))
