"""
JPMRGDPF File Generator
========================
Generates output DATA, META, and ZIP files.

This module handles:
- Creating DATA Excel files (Row 1: codes, Row 2: descriptions, Row 3+: data)
- Creating META Excel files with time series metadata
- Creating ZIP archives
- Updating the master data file
- Managing timestamped and latest output folders
"""

import os
import logging
import shutil
import zipfile

import pandas as pd
import openpyxl

import config

logger = logging.getLogger(__name__)


class JPMRGDPFFileGenerator:
    """Generates output files for JPMRGDPF runbook."""

    def __init__(self, column_order=None):
        self.column_order = column_order or config.COLUMN_ORDER
        self.timestamp = config.get_timestamp()

    def _create_output_dirs(self):
        """Create output directory structure."""
        ts_dir = os.path.join(config.OUTPUT_DIR, self.timestamp)
        os.makedirs(ts_dir, exist_ok=True)

        latest_dir = os.path.join(config.OUTPUT_DIR, config.LATEST_FOLDER)
        os.makedirs(latest_dir, exist_ok=True)

        os.makedirs(config.MASTER_DIR, exist_ok=True)

        return ts_dir, latest_dir

    def create_data_file(self, df, output_path):
        """
        Create DATA Excel file.

        Structure:
        - Row 1: Column codes (empty first cell, then codes)
        - Row 2: Column descriptions (empty first cell, then descriptions)
        - Row 3+: Year and data values

        Args:
            df: DataFrame with 'date' column + data columns.
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating DATA file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data'

        codes = [code for code, _ in self.column_order]
        descs = [desc for _, desc in self.column_order]

        # Row 1: codes
        ws.cell(row=1, column=1, value='')
        for i, code in enumerate(codes, start=2):
            ws.cell(row=1, column=i, value=code)

        # Row 2: descriptions
        ws.cell(row=2, column=1, value='')
        for i, desc in enumerate(descs, start=2):
            ws.cell(row=2, column=i, value=desc)

        # Row 3+: data
        for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
            # Date/year in first column - ensure clean integer year
            date_val = row['date']
            try:
                date_val = int(float(date_val))
            except (ValueError, TypeError):
                pass
            ws.cell(row=row_idx, column=1, value=date_val)

            for col_idx, code in enumerate(codes, start=2):
                if code in row:
                    value = row[code]
                    if pd.isna(value) if not isinstance(value, str) else False:
                        ws.cell(row=row_idx, column=col_idx, value=config.NA_VALUE)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=config.NA_VALUE)

        wb.save(output_path)
        wb.close()

        logger.info(f'DATA file created: {len(df)} rows, {len(codes)} columns')
        return output_path

    def create_meta_file(self, output_path):
        """
        Create META Excel file with time series metadata.

        Args:
            output_path: Path for output file.

        Returns:
            str: Path to created file.
        """
        logger.info(f'Creating META file: {output_path}')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Metadata'

        headers = ['CODE', 'DESCRIPTION', 'FREQUENCY', 'UNIT_TYPE', 'DATA_TYPE', 'SOURCE', 'CATEGORY']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        for row_idx, (code, desc) in enumerate(self.column_order, start=2):
            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=desc)
            ws.cell(row=row_idx, column=3, value=config.METADATA_DEFAULTS['FREQUENCY'])
            ws.cell(row=row_idx, column=4, value=config.METADATA_DEFAULTS['UNIT_TYPE'])
            ws.cell(row=row_idx, column=5, value=config.METADATA_DEFAULTS['DATA_TYPE'])
            ws.cell(row=row_idx, column=6, value=config.METADATA_DEFAULTS['SOURCE'])
            ws.cell(row=row_idx, column=7, value=config.METADATA_DEFAULTS['CATEGORY'])

        wb.save(output_path)
        wb.close()

        logger.info(f'META file created: {len(self.column_order)} entries')
        return output_path

    def create_zip_file(self, data_file, meta_file, zip_path):
        """
        Create ZIP archive containing DATA and META files.

        Args:
            data_file: Path to DATA file.
            meta_file: Path to META file.
            zip_path: Path for ZIP file.

        Returns:
            str: Path to created ZIP file.
        """
        logger.info(f'Creating ZIP file: {zip_path}')

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(data_file, os.path.basename(data_file))
            zf.write(meta_file, os.path.basename(meta_file))

        logger.info('ZIP file created')
        return zip_path

    def save_master_data(self, df):
        """
        Save updated master data file.

        Args:
            df: DataFrame with all accumulated data.

        Returns:
            str: Path to master file.
        """
        logger.info(f'Saving master data to: {config.MASTER_DATA_FILE}')
        self.create_data_file(df, config.MASTER_DATA_FILE)
        logger.info(f'Master data saved: {len(df)} rows')
        return config.MASTER_DATA_FILE

    def generate_files(self, df):
        """
        Generate all output files: timestamped + latest + master.

        Args:
            df: DataFrame with combined data.

        Returns:
            dict: Paths to all created files.
        """
        logger.info('Generating output files')

        ts_dir, latest_dir = self._create_output_dirs()

        # Timestamped file names
        data_name = f'{config.DATA_FILE_PREFIX}_{self.timestamp}.xlsx'
        meta_name = f'{config.META_FILE_PREFIX}_{self.timestamp}.xlsx'
        zip_name = f'{config.ZIP_FILE_PREFIX}_{self.timestamp}.zip'

        # Latest file names
        latest_data_name = f'{config.DATA_FILE_PREFIX}_LATEST.xlsx'
        latest_meta_name = f'{config.META_FILE_PREFIX}_LATEST.xlsx'
        latest_zip_name = f'{config.ZIP_FILE_PREFIX}_LATEST.zip'

        # Full paths
        ts_data = os.path.join(ts_dir, data_name)
        ts_meta = os.path.join(ts_dir, meta_name)
        ts_zip = os.path.join(ts_dir, zip_name)

        latest_data = os.path.join(latest_dir, latest_data_name)
        latest_meta = os.path.join(latest_dir, latest_meta_name)
        latest_zip = os.path.join(latest_dir, latest_zip_name)

        # Create timestamped files
        self.create_data_file(df, ts_data)
        self.create_meta_file(ts_meta)
        self.create_zip_file(ts_data, ts_meta, ts_zip)

        # Copy to latest
        shutil.copy2(ts_data, latest_data)
        shutil.copy2(ts_meta, latest_meta)
        shutil.copy2(ts_zip, latest_zip)
        logger.info('Copied files to latest folder')

        # Update master
        master_path = self.save_master_data(df)

        result = {
            'data_file': ts_data,
            'meta_file': ts_meta,
            'zip_file': ts_zip,
            'latest_data': latest_data,
            'latest_meta': latest_meta,
            'latest_zip': latest_zip,
            'master_file': master_path,
            'timestamp': self.timestamp,
            'output_dir': ts_dir,
        }

        logger.info('All output files generated successfully')
        return result
